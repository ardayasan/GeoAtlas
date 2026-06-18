#!/usr/bin/env python3
"""
Excel şablonu oluşturur: data/excel/template_nufus.xlsx
Örnek veriyle birlikte şablonu hazırlar.

Kullanım:
    python scripts/create_excel_template.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "excel", "template_nufus.xlsx")
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="EFF6FF")
THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row):
    for cell in ws[row]:
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')


def auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 22)


def build_province_sheet(wb):
    ws = wb.create_sheet("Nufus_Il")
    # il_kodu opsiyonel (il_adi'dan otomatik çözülür). Din alanları YOK.
    headers = [
        "il_kodu", "il_adi", "toplam_nufus", "erkek_nufus", "kadin_nufus",
        "yas_0_14", "yas_15_64", "yas_65_ust",
        "medyan_yas", "nufus_yogunluk", "nufus_artis_hizi", "veri_yili"
    ]
    ws.append(headers)
    style_header(ws)

    # Örnek: gerçek TÜİK 2025 değerleri (yaş grupları opsiyonel — boş bırakılabilir)
    examples = [
        ["", "İstanbul", 15754053, 7851344, 7902709, "", "", "", 35.0, 2943.4, 3.3, 2025],
        ["", "Ankara",    5864049, 2899485, 2964564, "", "", "", 34.4,  238.0, 9.2, 2025],
        ["", "İzmir",     4462056, 2210046, 2252010, "", "", "", 36.4,  372.0, 3.9, 2025],
        ["", "Sinop",      225848,  112000,  113848, "", "", "", 44.0,   40.0, -2.0, 2025],
        ["", "Şanlıurfa", 2265800, 1135000, 1130800, "", "", "", 21.8,  117.8, 6.0, 2025],
    ]
    for i, row in enumerate(examples, start=2):
        ws.append(row)
        style_data_row(ws, i)

    ws.row_dimensions[1].height = 36
    auto_col_width(ws)
    ws.freeze_panes = 'A2'
    return ws


def build_district_sheet(wb):
    ws = wb.create_sheet("Nufus_Ilce")
    # ilce_kodu opsiyonel (il_adi + ilce_adi'dan otomatik çözülür).
    headers = [
        "il_kodu", "il_adi", "ilce_kodu", "ilce_adi", "toplam_nufus",
        "erkek_nufus", "kadin_nufus",
        "yas_0_14", "yas_15_64", "yas_65_ust",
        "medyan_yas", "nufus_yogunluk", "veri_yili"
    ]
    ws.append(headers)
    style_header(ws)

    examples = [
        ["", "İstanbul", "", "Esenyurt", 1003905, 510615, 493290, "", "", "", "", "", 2025],
        ["", "İstanbul", "", "Kadıköy",   458638, 211000, 247638, "", "", "", "", "", 2025],
        ["", "Ankara",   "", "Çankaya",   952198, 455000, 497198, "", "", "", "", "", 2025],
    ]
    for i, row in enumerate(examples, start=2):
        ws.append(row)
        style_data_row(ws, i)

    ws.row_dimensions[1].height = 36
    auto_col_width(ws)
    ws.freeze_panes = 'A2'
    return ws


def build_labels_sheet(wb):
    ws = wb.create_sheet("Disaridan_Etiketler")
    headers = ["etiket_adi", "enlem", "boylam", "renk", "aciklama", "ikon"]
    ws.append(headers)
    style_header(ws)

    examples = [
        ["Galata Kulesi",  41.025610, 28.974080, "#FF5733", "İstanbul'un simgesi",          "pin"],
        ["Atatürk Havalimanı", 40.976922, 28.814606, "#3388FF", "Eski havalimanı alanı",    "flag"],
        ["Ankara Kalesi",  39.940000, 32.862000, "#9B59B6", "Tarihi kale",                  "star"],
    ]
    for i, row in enumerate(examples, start=2):
        ws.append(row)
        style_data_row(ws, i)

    # İkon seçenekleri notu
    ws.append([])
    ws.append(["* İkon seçenekleri:", "pin, star, circle, flag"])

    ws.row_dimensions[1].height = 36
    auto_col_width(ws)
    ws.freeze_panes = 'A2'
    return ws


def main():
    wb = openpyxl.Workbook()
    # Varsayılan sheet'i kaldır
    wb.remove(wb.active)

    build_province_sheet(wb)
    build_district_sheet(wb)
    build_labels_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"✓ Excel şablonu oluşturuldu: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
